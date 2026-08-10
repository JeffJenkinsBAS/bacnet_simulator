"""
Generates the BACnet Point-Mapping Worksheet directly from the live config
files (config/supervisory_device.json + config/devices/*.json), so it can
never drift from what the simulator actually publishes.
"""
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config_models import EquipmentGroupConfig, SupervisoryDeviceConfig, validate_equipment_groups

PROJECT_DIR = Path(__file__).resolve().parent.parent
CONFIG_DIR = PROJECT_DIR / "config"
OUT_PATH = PROJECT_DIR / "ACI_BACnet_Simulator_Point_Mapping.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
INTERLOCK_FILL = PatternFill("solid", fgColor="FDE2E1")
SIM_TO_WEBCTRL_FONT = Font(name=FONT_NAME, size=10, color="15803D")
WEBCTRL_TO_SIM_FONT = Font(name=FONT_NAME, size=10, color="1D4ED8")
THIN_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))

OBJECT_TYPE_ABBR = {
    "analog-input": "AI", "analog-output": "AO", "analog-value": "AV",
    "binary-input": "BI", "binary-output": "BO", "binary-value": "BV",
    "multi-state-input": "MSI", "multi-state-output": "MSO", "multi-state-value": "MSV",
}


def load_supervisory() -> SupervisoryDeviceConfig:
    with open(CONFIG_DIR / "supervisory_device.json") as f:
        return SupervisoryDeviceConfig.model_validate(json.load(f))


def load_network():
    with open(CONFIG_DIR / "network.json") as f:
        return json.load(f)


def load_groups() -> list[EquipmentGroupConfig]:
    groups = []
    for path in sorted((CONFIG_DIR / "devices").glob("*.json")):
        with open(path) as f:
            groups.append(EquipmentGroupConfig.model_validate(json.load(f)))
    validate_equipment_groups(groups)
    return sorted(groups, key=lambda g: g.instance_offset)


def style_header_row(ws, row_num, headers):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row_num + 1, column=1).coordinate


def autosize(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_workbook():
    supervisory = load_supervisory()
    network = load_network()
    groups = load_groups()

    wb = Workbook()

    # ---- Sheet 1: Device Summary ----------------------------------------
    ws = wb.active
    ws.title = "Device Summary"
    ws["A1"] = "ACI BACnet Building Simulation Platform — Device Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Generated directly from the running project's config files — see Point List tab for every object."
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    row = 4
    device_fields = [
        ("Device Name", supervisory.device_name),
        ("Device Instance", supervisory.device_instance),
        ("Description", supervisory.description),
        ("Bind Address", network["bind_address"] + " (set to the bench NIC's real IP before connecting -- never 0.0.0.0)"),
        ("UDP Port", network["udp_port"]),
        ("Vendor Identifier", network["vendor_identifier"]),
        ("Private Lab Mode", network["private_lab_mode"]),
        ("Respond to Who-Is", network["respond_to_who_is"]),
    ]
    for label, value in device_fields:
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT_NAME, bold=True, size=10)
        ws.cell(row=row, column=2, value=str(value)).font = NORMAL_FONT
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="This single BACnet device hosts every equipment group's objects below.").font = SUBTITLE_FONT
    row += 2

    headers = ["Equipment Group", "Instance Offset", "Point Count", "Description"]
    style_header_row(ws, row, headers)
    row += 1
    total_points = 0
    for g in groups:
        ws.cell(row=row, column=1, value=g.group_id).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=g.instance_offset).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=len(g.points)).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=g.description).font = NORMAL_FONT
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        total_points += len(g.points)
        row += 1

    ws.cell(row=row + 1, column=1, value="TOTAL").font = Font(name=FONT_NAME, bold=True, size=10)
    ws.cell(row=row + 1, column=2, value=len(groups)).font = Font(name=FONT_NAME, bold=True, size=10)
    ws.cell(row=row + 1, column=3, value=total_points).font = Font(name=FONT_NAME, bold=True, size=10)

    autosize(ws, [26, 16, 12, 90])

    # ---- Sheet 2: Point List (the actual matrix) ------------------------
    ws2 = wb.create_sheet("Point List")
    ws2["A1"] = "ACI BACnet Building Simulation Platform — Full Point-Mapping Matrix"
    ws2["A1"].font = TITLE_FONT
    ws2["A2"] = (
        f"All points live under ONE BACnet device: {supervisory.device_name}, instance {supervisory.device_instance}, "
        f"port {network['udp_port']}. 'Object Instance' below is the real, global instance on the wire "
        f"(= equipment group's instance offset + local instance)."
    )
    ws2["A2"].font = SUBTITLE_FONT
    ws2.merge_cells("A1:N1")
    ws2.merge_cells("A2:N2")

    headers = [
        "Equipment Group", "Point Alias", "BACnet Object Name", "Object Type", "Object Instance",
        "BACnet Address String", "Signal Direction", "Units", "Writable", "Interlock",
        "Initial Value", "Normal Range Low", "Normal Range High", "Description",
    ]
    header_row = 4
    style_header_row(ws2, header_row, headers)

    r = header_row + 1
    for g in groups:
        for p in g.points:
            global_instance = g.instance_offset + p.object_instance
            abbr = OBJECT_TYPE_ABBR[p.object_type.value]
            address_string = f"bacnet://{supervisory.device_instance}/{abbr.lower()}:{global_instance}"

            values = [
                g.group_id,
                p.alias,
                p.object_name,
                f"{p.object_type.value} ({abbr})",
                global_instance,
                address_string,
                p.signal_direction.value,
                p.units,
                "Yes" if p.writable else "No",
                "YES -- HARD INTERLOCK" if p.interlock else "",
                p.initial_value,
                p.normal_range.low if p.normal_range else "",
                p.normal_range.high if p.normal_range else "",
                p.description,
            ]
            for c, v in enumerate(values, start=1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                if c == 7:  # Signal Direction column
                    if p.signal_direction.value == "sim_to_webctrl":
                        cell.font = SIM_TO_WEBCTRL_FONT
                    elif p.signal_direction.value == "webctrl_to_sim":
                        cell.font = WEBCTRL_TO_SIM_FONT
                if p.interlock:
                    cell.fill = INTERLOCK_FILL
            r += 1

    ws2.auto_filter.ref = f"A{header_row}:N{r-1}"
    autosize(ws2, [20, 26, 32, 16, 14, 34, 16, 22, 9, 20, 12, 14, 14, 60])

    # ---- Sheet 3: Legend -------------------------------------------------
    ws3 = wb.create_sheet("Legend")
    ws3["A1"] = "Legend"
    ws3["A1"].font = TITLE_FONT
    legend_rows = [
        ("Object Type abbreviations", "AI=analog-input, AO=analog-output, AV=analog-value, "
                                       "BI=binary-input, BO=binary-output, BV=binary-value "
                                       "(no multi-state points exist in this simulator yet)"),
        ("Signal Direction: sim_to_webctrl", "Simulator publishes this value; WebCTRL reads it. Shown in green."),
        ("Signal Direction: webctrl_to_sim", "WebCTRL/EIKON writes this value; the simulator's equipment model reacts to it. Shown in blue."),
        ("Writable", "Whether this object accepts BACnet WriteProperty (has a priority array)."),
        ("Interlock", "A hard interlock -- checked first every simulation tick, ahead of normal commands, "
                       "and forces the equipment into a fixed safe state while active. Real physical safety "
                       "devices on the bench (Freezestat, High Static Pressure, Emerg/Refrig Shutdown) relay "
                       "into these via real BACnet writes. Highlighted in red."),
        ("BACnet Address String", "Example EIKON Network I/O microblock address format, per the ALC BACnet "
                                    "Integration Guide: bacnet://<device instance>/<object type>:<object instance>[/property][@priority]. "
                                    "Add a property (e.g. /priority-array) or priority (e.g. @8) as needed -- see Appendix A of that guide."),
        ("Instance numbering scheme", "Each equipment group's instance_offset (see Device Summary tab) plus its point's "
                                        "small local instance number gives the real global Object Instance shown here. "
                                        "Offsets are (group's position in the fleet) x 1000, e.g. AHU-1 = 9000."),
        ("Zone Temp for VAV-1 / VAV-2", "Not simulated -- sourced from real communicating ZS thermostats on the bench, "
                                          "per Jeff's confirmation. VAV-3 through VAV-17 are virtual zones and DO publish a simulated Zone Temp."),
        ("Source of truth", "This workbook is generated directly from config/supervisory_device.json and "
                              "config/devices/*.json via scripts/generate_point_mapping_workbook.py -- re-run that "
                              "script any time the object model changes rather than hand-editing this file."),
    ]
    row = 3
    for label, text in legend_rows:
        ws3.cell(row=row, column=1, value=label).font = Font(name=FONT_NAME, bold=True, size=10)
        ws3.cell(row=row, column=2, value=text).font = NORMAL_FONT
        ws3.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[row].height = 30
        row += 1
    autosize(ws3, [30, 110])

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"{len(groups)} equipment groups, {total_points} total points")


if __name__ == "__main__":
    build_workbook()
